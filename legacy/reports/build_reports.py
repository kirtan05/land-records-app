#!/usr/bin/env python3
"""Per-survey combined PDF (clickable index + bookmarks) + master Excel with hyperlinks."""
import os, json, datetime
from io import BytesIO
from fpdf import FPDF
from pypdf import PdfReader, PdfWriter
from pypdf.annotations import Link
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import sys, os; sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '../../tools'))
from repo_root import REPO

BASE = str(REPO)
OUT = os.path.join(BASE, "output")
FONTS = "/usr/share/fonts/noto"
K = 72 / 25.4                      # mm -> pt
PAGE_H_PT = 297 * K
TODAY = datetime.date.today().isoformat()

state = json.load(open(os.path.join(OUT, "_state.json")))["surveys"]

# ---------- index page (fpdf2) ----------
COLS = [("#", 12, 9), ("Registration No", 21, 59), ("Status", 80, 23),
        ("Parties (Petitioner  v/s  Respondent)", 103, 83), ("Pg", 186, 12)]
RH = 8.0

def new_pdf():
    p = FPDF(orientation="P", unit="mm", format="A4")
    p.add_font("noto", "", f"{FONTS}/NotoSans-Regular.ttf")
    p.add_font("noto", "B", f"{FONTS}/NotoSans-Bold.ttf")
    p.add_font("guj", "", f"{FONTS}/NotoSansGujarati-Regular.ttf")
    p.add_font("guj", "B", f"{FONTS}/NotoSansGujarati-Bold.ttf")
    p.set_fallback_fonts(["guj"])
    p.set_auto_page_break(False)
    return p

def fit(p, s, w):
    s = (s or "").replace("\n", " ")
    if p.get_string_width(s) <= w:
        return s
    while s and p.get_string_width(s + "…") > w:
        s = s[:-1]
    return s + "…"

def header(p):
    p.set_xy(12, p.get_y())
    p.set_font("noto", "B", 9)
    p.set_fill_color(31, 78, 120); p.set_text_color(255)
    for name, x, w in COLS:
        p.set_xy(x, p.get_y()); p.cell(w, 7, " " + fit(p, name, w - 2), border=0, fill=True)
    p.set_xy(12, p.get_y() + 7); p.set_text_color(0)

def make_index(key, cases, starts, disposed, pending):
    p = new_pdf(); p.add_page()
    p.set_font("noto", "B", 16); p.set_text_color(31, 78, 120)
    p.set_xy(12, 13); p.cell(0, 8, f"Survey No {key}  -  Village Bharoda (Anand / Umreth)")
    p.set_font("noto", "", 10); p.set_text_color(90)
    p.set_xy(12, 22)
    p.cell(0, 6, f"{len(cases)} cases  -  {disposed} disposed, {pending} pending   |   generated {TODAY}   |   click a row to jump; use the Bookmarks panel too")
    p.set_text_color(0); p.set_xy(12, 31); header(p)
    rects = []
    for i, c in enumerate(cases):
        y = p.get_y()
        if y + RH > 283:
            p.add_page(); p.set_xy(12, 14); header(p); y = p.get_y()
        rec = c["rec"]
        if i % 2 == 0:
            p.set_fill_color(244, 247, 250)
            p.set_xy(12, y); p.cell(186, RH, "", fill=True)
        p.set_font("noto", "", 9)
        p.set_xy(COLS[0][1], y); p.cell(COLS[0][2], RH, str(rec["case_index"]))
        p.set_xy(COLS[1][1], y); p.cell(COLS[1][2], RH, " " + fit(p, rec.get("registration_no") or rec.get("case_str", ""), COLS[1][2] - 2))
        st = (rec.get("status") or "").upper()
        p.set_text_color(*( (0, 120, 60) if "DISPOS" in st else (170, 90, 0)))
        p.set_font("noto", "B", 9)
        p.set_xy(COLS[2][1], y); p.cell(COLS[2][2], RH, " " + fit(p, st or "-", COLS[2][2] - 2))
        p.set_text_color(0); p.set_font("noto", "", 9)
        p.set_xy(COLS[3][1], y); p.cell(COLS[3][2], RH, " " + fit(p, rec.get("parties", ""), COLS[3][2] - 2))
        p.set_xy(COLS[4][1], y); p.cell(COLS[4][2], RH, str(starts[i] + 1))
        p.set_xy(12, y + RH)
        rect = (12 * K, PAGE_H_PT - (y + RH) * K, 198 * K, PAGE_H_PT - y * K)
        rects.append((p.page - 1, rect, starts[i]))
    return bytes(p.output()), p.page, rects

# ---------- build per-survey combined PDF ----------
def build_pdf(key, token, records):
    cases = []
    for rec in sorted(records, key=lambda r: r["case_index"]):
        srcs = []
        cp = rec.get("case_pdf")
        if cp and os.path.exists(cp):
            srcs.append(("case", PdfReader(cp)))
        for op in str(rec.get("order_pdf", "")).split(";"):
            op = op.strip()
            if op and os.path.exists(op):
                srcs.append(("order", PdfReader(op)))
        cases.append({"rec": rec, "srcs": srcs, "pages": sum(len(r.pages) for _, r in srcs)})

    disposed = sum(1 for c in cases if "DISPOS" in (c["rec"].get("status") or "").upper())
    pending = len(cases) - disposed
    idx_n = 1
    for _ in range(4):
        starts, acc = [], idx_n
        for c in cases:
            starts.append(acc); acc += c["pages"]
        index_bytes, pc, rects = make_index(key, cases, starts, disposed, pending)
        if pc == idx_n:
            break
        idx_n = pc

    w = PdfWriter()
    for pg in PdfReader(BytesIO(index_bytes)).pages:
        w.add_page(pg)
    w.add_outline_item("Index", 0)
    for ci, c in enumerate(cases):
        rec = c["rec"]
        title = f"Case {rec['case_index']}: {rec.get('registration_no') or rec.get('case_str','')[:40]}  [{rec.get('status','')}]"
        parent = w.add_outline_item(title, starts[ci])
        first_order = True
        for kind, r in c["srcs"]:
            if kind == "order" and first_order:
                w.add_outline_item("  ↳ Order / Judgement", len(w.pages), parent=parent)
                first_order = False
            for pg in r.pages:
                w.add_page(pg)
    for page_no, rect, target in rects:
        w.add_annotation(page_number=page_no, annotation=Link(rect=rect, target_page_index=target))

    out_path = os.path.join(OUT, token, f"Bharoda_SurveyNo_{token}_ALL.pdf")
    with open(out_path, "wb") as f:
        w.write(f)
    return out_path, len(w.pages)

# ---------- master Excel ----------
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(color="FFFFFF", bold=True, size=12)
GREEN = PatternFill("solid", fgColor="C6EFCE"); AMBER = PatternFill("solid", fgColor="FFEB9C")
LINK = Font(color="0563C1", underline="single", size=11)
THIN = Border(*[Side(style="thin", color="D9D9D9")] * 4)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")

def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    ws.row_dimensions[row].height = 30

def linkcell(ws, r, c, target, text="📄 Open"):
    cell = ws.cell(row=r, column=c, value=text if target and os.path.exists(target) else "—")
    if target and os.path.exists(target):
        cell.hyperlink = target; cell.font = LINK
    cell.alignment = TOP

def build_excel():
    wb = Workbook()
    # ---- All Cases sheet ----
    ws = wb.active; ws.title = "All Cases"
    cols = ["Survey No", "Case #", "Type", "Registration No", "Status", "Disposal Date",
            "Parties (Petitioner v/s Respondent)", "Office", "Order?", "Case PDF", "Order PDF", "Combined (survey)"]
    ws.append(cols); style_header(ws, len(cols))
    r = 2
    for key, info in state.items():
        if info.get("status") != "done":
            continue
        token = info["token"]
        summ = json.load(open(os.path.join(OUT, token, "_summary.json")))
        allpdf = os.path.join(OUT, token, f"Bharoda_SurveyNo_{token}_ALL.pdf")
        for rec in sorted(summ["records"], key=lambda x: x["case_index"]):
            reg = rec.get("registration_no") or rec.get("case_str", "")
            typ = reg.split("/")[0].strip()
            ws.cell(row=r, column=1, value=key).alignment = TOP
            ws.cell(row=r, column=2, value=rec["case_index"]).alignment = TOP
            ws.cell(row=r, column=3, value=typ).alignment = TOP
            ws.cell(row=r, column=4, value=reg).alignment = TOP
            st = (rec.get("status") or "").upper()
            sc = ws.cell(row=r, column=5, value=st); sc.alignment = TOP
            sc.fill = GREEN if "DISPOS" in st else AMBER
            dd = (rec.get("disposal_date", "") or "").split("Disposal Type")[0].strip()
            ws.cell(row=r, column=6, value=dd).alignment = TOP
            pc = ws.cell(row=r, column=7, value=rec.get("parties", "")); pc.alignment = WRAP
            ws.cell(row=r, column=8, value=rec.get("office", "")).alignment = WRAP
            ws.cell(row=r, column=9, value="Yes" if rec.get("order_downloaded") else "No").alignment = TOP
            linkcell(ws, r, 10, rec.get("case_pdf"))
            ords = [o for o in str(rec.get("order_pdf", "")).split(";") if o.strip()]
            linkcell(ws, r, 11, ords[0] if ords else None)
            linkcell(ws, r, 12, allpdf, "📚 Open")
            r += 1
    widths = [11, 7, 15, 30, 11, 13, 60, 20, 8, 12, 12, 16]
    for i, wd in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = wd
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:L{r-1}"

    # ---- Summary sheet (iRCMS case counts + AnyRoR land record + old-scanned VF-7/12, per survey) ----
    ap = os.path.join(OUT, "_anyror_state.json")
    anyror = json.load(open(ap)) if os.path.exists(ap) else {}
    vp = os.path.join(OUT, "_vf712_state.json")
    vf712 = json.load(open(vp)) if os.path.exists(vp) else {}
    def assess(info):
        jp = info.get("json")
        if jp and os.path.exists(jp):
            try: return json.load(open(jp)).get("total_assessment", "") or ""
            except Exception: return ""
        return ""

    s = wb.create_sheet("Summary", 0)
    scols = ["Survey No", "Cases", "Disposed", "Pending", "Orders", "Combined PDF (iRCMS)", "Folder",
             "Total Area (H-Are-SqMt)", "Assessment", "Tenure", "Land Use", "Land Record (AnyRoR)",
             "VF-7/12 Docs", "VF-7/12 (old scanned)"]
    s.append(scols); style_header(s, len(scols))
    rr = 2; tot = [0, 0, 0, 0]
    for key, info in state.items():
        token = info.get("token")
        if info.get("status") == "done":
            summ = json.load(open(os.path.join(OUT, token, "_summary.json")))
            recs = summ["records"]
            disp = sum(1 for x in recs if "DISPOS" in (x.get("status") or "").upper())
            orders = sum(1 for x in recs if x.get("order_downloaded"))
            vals = [key, len(recs), disp, len(recs) - disp, orders]
            tot = [tot[0] + len(recs), tot[1] + disp, tot[2] + (len(recs) - disp), tot[3] + orders]
            for c, v in enumerate(vals, 1):
                s.cell(row=rr, column=c, value=v).alignment = TOP
            linkcell(s, rr, 6, os.path.join(OUT, token, f"Bharoda_SurveyNo_{token}_ALL.pdf"), "📚 Open")
            linkcell(s, rr, 7, os.path.join(OUT, token), "📂 Open")
        else:
            for c, v in enumerate([key, 0, 0, 0, 0], 1):
                s.cell(row=rr, column=c, value=v).alignment = TOP
            s.cell(row=rr, column=6, value="(no iRCMS cases)").font = Font(italic=True, color="888888")
        # AnyRoR land-record columns (per survey)
        ar = anyror.get(key)
        if ar and ar.get("done"):
            s.cell(row=rr, column=8, value=ar.get("area", "")).alignment = TOP
            s.cell(row=rr, column=9, value=assess(ar)).alignment = TOP
            s.cell(row=rr, column=10, value=ar.get("tenure", "")).alignment = WRAP
            s.cell(row=rr, column=11, value=ar.get("land_use", "")).alignment = WRAP
            linkcell(s, rr, 12, ar.get("pdf"), "📄 Open")
        # VF-7/12 old-scanned columns (per survey)
        vf = vf712.get(key)
        if vf and vf.get("kept"):
            s.cell(row=rr, column=13, value=vf.get("kept")).alignment = TOP
            linkcell(s, rr, 14, vf.get("combined"), "📄 Open")
        rr += 1
    s.cell(row=rr, column=1, value="TOTAL").font = Font(bold=True)
    for c, v in zip([2, 3, 4, 5], tot):
        s.cell(row=rr, column=c, value=v).font = Font(bold=True)
    for i, wd in enumerate([12, 8, 10, 9, 8, 18, 9, 18, 12, 22, 22, 18, 11, 18], 1):
        s.column_dimensions[chr(64 + i)].width = wd
    s.freeze_panes = "A2"; s.auto_filter.ref = f"A1:N{rr-1}"

    path = os.path.join(OUT, "iRCMS_Bharoda_Master.xlsx")
    wb.save(path)
    return path

# ---------- run ----------
done = 0
for key, info in state.items():
    if info.get("status") != "done":
        print(f"  {key}: {info.get('status')} (skip pdf)"); continue
    token = info["token"]
    summ = json.load(open(os.path.join(OUT, token, "_summary.json")))
    path, pages = build_pdf(key, token, summ["records"])
    print(f"  ✓ {key}: {os.path.basename(path)} ({pages} pages, {len(summ['records'])} cases)")
    done += 1
xlsx = build_excel()
print(f"\nbuilt {done} combined PDFs + master Excel -> {xlsx}")
